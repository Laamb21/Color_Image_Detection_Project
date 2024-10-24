# gui.py

import os
import sys
import subprocess
import threading
import queue
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, scrolledtext, messagebox

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from processing import process_documents

import csv  # Needed for parsing the TSV log file
import shutil  # Needed for copying files

import logging  # Ensure logging is imported if not already

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("JPG and TIFF Processor")
        self.root.geometry("800x700")
        
        # Initialize variables
        self.parent_folder = tk.StringVar()
        self.low_threshold = tk.DoubleVar(value=10.0)   # Default low threshold
        self.high_threshold = tk.DoubleVar(value=15.0)  # Default high threshold
        self.processing_thread = None
        self.progress_queue = queue.Queue()
        
        # List to keep track of flagged files
        self.flagged_files = []

        # List to keep track of selected files 
        self.selected_files = []
        
        # Initialize log entries
        self.log_entries = []  # To store log data
        
        # Create GUI components
        self.create_widgets()
    
    def create_widgets(self):
        # ---------------------------- Folder Selection ---------------------------- #
        folder_frame = ttk.Frame(self.root)
        folder_frame.pack(padx=10, pady=10, fill='x')
        
        ttk.Label(folder_frame, text="Parent Folder:").pack(side='left', padx=(0,5))
        ttk.Entry(folder_frame, textvariable=self.parent_folder, width=50).pack(side='left', fill='x', expand=True)
        ttk.Button(folder_frame, text="Select Folder", command=self.select_folder).pack(side='left', padx=(5,0))
        
        # ---------------------------- Thresholds ---------------------------- #
        threshold_frame = ttk.Frame(self.root)
        threshold_frame.pack(padx=10, pady=5, fill='x')
        
        ttk.Label(threshold_frame, text="Low Gray Threshold (%):").pack(side='left', padx=(0,5))
        # Adding validation to ensure only numeric input
        vcmd = (self.root.register(self.validate_threshold), '%P')
        ttk.Entry(threshold_frame, textvariable=self.low_threshold, width=10, validate='key', validatecommand=vcmd).pack(side='left', padx=(0,15))
        
        ttk.Label(threshold_frame, text="High Gray Threshold (%):").pack(side='left', padx=(0,5))
        ttk.Entry(threshold_frame, textvariable=self.high_threshold, width=10, validate='key', validatecommand=vcmd).pack(side='left')
        
        # ---------------------------- Run Button ---------------------------- #
        run_frame = ttk.Frame(self.root)
        run_frame.pack(padx=10, pady=10, fill='x')
        
        self.run_button = ttk.Button(run_frame, text="Run Script", command=self.run_script)
        self.run_button.pack(side='left')
        
        # ---------------------------- Progress Bar ---------------------------- #
        progress_frame = ttk.Frame(self.root)
        progress_frame.pack(padx=10, pady=10, fill='x')
        
        self.progress = ttk.Progressbar(progress_frame, orient='horizontal', mode='determinate')
        self.progress.pack(fill='x', expand=True)
        
        # Adding a label to show percentage
        self.progress_label = ttk.Label(progress_frame, text="0%")
        self.progress_label.pack(pady=(5,0))
        
        # ---------------------------- Log Display ---------------------------- #
        log_frame = ttk.Frame(self.root)
        log_frame.pack(padx=10, pady=10, fill='both', expand=True)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, state='disabled')
        self.log_text.pack(fill='both', expand=True)
        
        # ---------------------------- Download Log Buttons ---------------------------- #
        buttons_frame = ttk.Frame(self.root)
        buttons_frame.pack(padx=10, pady=10, fill='x')
        
        # Renamed from "Open TSV Log" to "Download TSV Log"
        self.download_tsv_button = ttk.Button(buttons_frame, text="Download TSV Log", command=self.download_tsv_log, state='disabled')
        self.download_tsv_button.pack(side='left', padx=(0,5))
        
        # Renamed from "Open Excel Log" to "Download Excel Log"
        self.download_excel_button = ttk.Button(buttons_frame, text="Download Excel Log", command=self.download_excel_log, state='disabled')
        self.download_excel_button.pack(side='left')
        
        # ---------------------------- Download Flagged Files Button ---------------------------- #
        self.download_flagged_button = ttk.Button(buttons_frame, text="Download Flagged Files", command=self.download_flagged_files, state='disabled')
        self.download_flagged_button.pack(side='left', padx=(5,0))

        # ---------------------------- Download Selected Files Button ---------------------------- #
        self.download_selected_button = ttk.Button(buttons_frame, text="Download Selected Files", command=self.download_selected_files, state='disabled')
        self.download_selected_button.pack(side='left', padx=(5,0))

    def validate_threshold(self, P):
        """Validate that the input is a valid float."""
        if P == "":
            return True
        try:
            float(P)
            return True
        except ValueError:
            return False
    
    def select_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.parent_folder.set(folder_selected)
    
    def run_script(self):
        if not self.parent_folder.get():
            messagebox.showerror("Error", "Please select a parent folder.")
            return
        
        # Validate that low_threshold is less than high_threshold
        if self.low_threshold.get() >= self.high_threshold.get():
            messagebox.showerror("Error", "Low threshold must be less than High threshold.")
            return
        
        # Disable the Run button to prevent multiple runs
        self.run_button.config(state='disabled')
        
        # Clear previous logs
        self.log_text.configure(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.configure(state='disabled')
        
        # Initialize progress
        self.progress['value'] = 0
        self.progress['maximum'] = 100  # Will be updated dynamically
        self.progress_label.config(text="0%")
        
        # Reset flagged files list
        self.flagged_files = []
        self.download_flagged_button.config(state='disabled')  # Disable until processing is complete
        
        # Disable download buttons until processing is complete
        self.download_tsv_button.config(state='disabled')
        self.download_excel_button.config(state='disabled')
        
        # Clear previous log entries
        self.log_entries = []
        
        # Start the processing in a separate thread
        self.processing_thread = threading.Thread(
            target=self.process,
            args=()
        )
        self.processing_thread.start()
        
        # Start polling the queue
        self.root.after(100, self.process_queue)
    
    def process(self):
        input_dir_jpg = os.path.join(self.parent_folder.get(), "JPG")
        input_dir_tiff = os.path.join(self.parent_folder.get(), "TIF")
        
        # Call the processing function without specifying log file paths
        process_documents(
            input_dir_jpg,
            input_dir_tiff,
            self.progress_queue,
            self.low_threshold.get(),
            self.high_threshold.get()
        )
    
    def process_queue(self):
        try:
            while True:
                message = self.progress_queue.get_nowait()
                if message[0] == "current_file":
                    self.log_text.configure(state='normal')
                    self.log_text.insert(tk.END, f"Processing: {message[1]}\n")
                    self.log_text.configure(state='disabled')
                elif message[0] == "progress":
                    processed, total = message[1], message[2]
                    if total > 0:
                        progress_value = (processed / total) * 100
                        self.progress['value'] = progress_value
                        self.progress_label.config(text=f"{progress_value:.2f}%")
                elif message[0] == "error":
                    self.log_text.configure(state='normal')
                    self.log_text.insert(tk.END, f"Error: {message[1]}\n")
                    self.log_text.configure(state='disabled')
                    messagebox.showerror("Error", message[1])
                elif message[0] == "complete":
                    completion_message = message[1]
                    flagged_count = message[2]
                    log_entries_sorted = message[3]  # Retrieve log data
                    
                    self.log_text.configure(state='normal')
                    self.log_text.insert(tk.END, f"{completion_message}\nFlagged Files Count: {flagged_count}\n")
                    self.log_text.configure(state='disabled')
                    messagebox.showinfo("Complete", f"{completion_message}\nFlagged Files Count: {flagged_count}")
                    
                    # Store the log entries for downloading
                    self.log_entries = log_entries_sorted
                    
                    # Enable the Download Log Buttons
                    self.download_tsv_button.config(state='normal')
                    self.download_excel_button.config(state='normal')
                    
                    # Enable the Download Flagged Files Button
                    self.download_flagged_button.config(state='normal')

                    #Enable the Download Selected Files Button
                    self.download_selected_button.config(state='normal')
                    
                    # Re-enable the Run button
                    self.run_button.config(state='normal')
                    
                    # Populate the flagged_files list by reading the log entries
                    self.populate_flagged_files()

                    # Populate the selected files list based on log entries
                    self.populate_selected_files()
                    
        except queue.Empty:
            pass
        if self.processing_thread and self.processing_thread.is_alive():
            self.root.after(100, self.process_queue)
    
    def populate_flagged_files(self):
        """Reads the log entries and populates the flagged_files list."""
        input_dir_tiff = os.path.join(self.parent_folder.get(), "TIF")
        
        for entry in self.log_entries:
            sort_key, selected_documents, gray_pct_str, selected_format, flagged = entry
            if flagged == "Yes":
                documents = selected_documents.split(', ')
                for document in documents:
                    # Reconstruct the TIFF filename (assuming .tif or .tiff extension)
                    possible_extensions = ['.tif', '.tiff']
                    for ext in possible_extensions:
                        filename = document + ext
                        filepath = os.path.join(input_dir_tiff, filename)
                        if os.path.exists(filepath):
                            self.flagged_files.append(filepath)
                            break
                    else:
                        # If none of the extensions matched, notify the user
                        self.log_text.configure(state='normal')
                        self.log_text.insert(tk.END, f"Flagged file '{document}' not found with extensions .tif or .tiff.\n")
                        self.log_text.configure(state='disabled')
        if not self.flagged_files:
            self.log_text.configure(state='normal')
            self.log_text.insert(tk.END, "No flagged files found.\n")
            self.log_text.configure(state='disabled')
    
    def populate_selected_files(self):
        """Populates the selected_files list based on the selected_format in log_entries."""
        input_dir_jpg = os.path.join(self.parent_folder.get(), "JPG")
        input_dir_tiff = os.path.join(self.parent_folder.get(), "TIF")
        
        for entry in self.log_entries:
            sort_key, selected_documents, gray_pct_str, selected_format, flagged = entry
            documents = selected_documents.split(', ')
            for document in documents:
                # Depending on the selected format,, determine the file type
                if selected_format == "JPG":
                    # Assume JPG files have .jpg or .jpeg extensions
                    possible_extensions = ['.jpg', '.jpeg']
                    for ext in possible_extensions:
                        filename = document + ext
                        filepath = os.path.join(input_dir_jpg, filename)
                        if os.path.exists(filepath):
                            self.selected_files.append(filepath)
                            break
                elif selected_format in ["TIFF", "TIF (Intermediate)"]:
                    # Assume TIFF files have .tif or .tiff extensions
                    possible_extensions = ['.tif', '.tiff']
                    for ext in possible_extensions:
                        filename = document + ext
                        filepath = os.path.join(input_dir_tiff, filename)
                        if os.path.exists(filepath):
                            self.selected_files.append(filepath)
                            break
                # If needed, handle other formats
                    
        if not self.selected_files:
            self.log_text.configure(state='normal')
            self.log_text.insert(tk.END, "No selected files found.\n")
            self.log_text.configure(state='disabled')


    def download_tsv_log(self):
        """Handles downloading the TSV log file to a user-selected location."""
        if not self.log_entries:
            messagebox.showerror("Error", "No log data available to download.")
            return
        
        # Prompt user to select the destination file path
        destination_path = filedialog.asksaveasfilename(
            defaultextension=".tsv",
            filetypes=[("TSV files", "*.tsv"), ("All files", "*.*")],
            title="Save TSV Log As"
        )
        if destination_path:
            try:
                with open(destination_path, mode='w', newline='') as log_csv:
                    log_writer = csv.writer(log_csv, delimiter='\t')  # Tab delimiter
                    # Write header with the fifth column
                    log_writer.writerow(['Document', 'Gray_Percentage', 'Selected_Format', 'Flagged_Files'])
    
                    # Write data rows
                    for entry in self.log_entries:
                        _, selected_documents, gray_pct_str, selected_format, flagged = entry
                        log_writer.writerow([selected_documents, gray_pct_str, selected_format, flagged])
                messagebox.showinfo("Success", f"TSV log has been downloaded to:\n{destination_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to download TSV log: {str(e)}")
    
    def download_excel_log(self):
        """Handles downloading the Excel log file to a user-selected location."""
        if not self.log_entries:
            messagebox.showerror("Error", "No log data available to download.")
            return
        
        # Prompt user to select the destination file path
        destination_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Excel Log As"
        )
        if destination_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Selection Log"
    
                # Define headers
                headers = ['Document', 'Gray Percentage (%)', 'Selected Format', 'Flagged_Files']
                header_font = Font(bold=True)
    
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = header_font
    
                # Write data rows
                for row_num, entry in enumerate(self.log_entries, start=2):
                    _, selected_documents, gray_pct_str, selected_format, flagged = entry
                    ws.cell(row=row_num, column=1, value=selected_documents)
                    ws.cell(row=row_num, column=2, value=float(gray_pct_str))
                    ws.cell(row=row_num, column=3, value=selected_format)
                    ws.cell(row=row_num, column=4, value=flagged)
    
                # Adjust column widths for better readability
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
    
                # Save the workbook
                wb.save(destination_path)
                messagebox.showinfo("Success", f"Excel log has been downloaded to:\n{destination_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to download Excel log: {str(e)}")
    
    def download_flagged_files(self):
        if not self.flagged_files:
            messagebox.showinfo("No Flagged Files", "There are no flagged files to download.")
            return
        
        # Prompt user to select the destination folder where "Flagged Files" will be created
        destination_parent = filedialog.askdirectory(title="Select Destination for 'Flagged Files' Folder")
        if not destination_parent:
            return  # User cancelled the dialog
        
        # Define the "Flagged Files" subfolder path
        flagged_folder_name = "Flagged Files"
        destination_folder = os.path.join(destination_parent, flagged_folder_name)
        
        # Check if "Flagged Files" folder already exists
        if not os.path.exists(destination_folder):
            try:
                os.makedirs(destination_folder)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create 'Flagged Files' folder: {str(e)}")
                return
        else:
            # Ask the user if they want to overwrite existing files
            overwrite = messagebox.askyesno("Folder Exists", f"The folder '{flagged_folder_name}' already exists in the selected destination.\nDo you want to overwrite its contents?")
            if overwrite:
                # Optionally, clear existing contents
                try:
                    for filename in os.listdir(destination_folder):
                        file_path = os.path.join(destination_folder, filename)
                        if os.path.isfile(file_path):
                            os.remove(file_path)
                        elif os.path.isdir(file_path):
                            shutil.rmtree(file_path)
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to clear existing 'Flagged Files' folder: {str(e)}")
                    return
            else:
                return  # User chose not to overwrite; cancel the operation
        
        # Start copying in a separate thread to keep GUI responsive
        copy_thread = threading.Thread(target=self.copy_flagged_files, args=(destination_folder,))
        copy_thread.start()

    def download_selected_files(self):
        if not self.selected_files:
            messagebox.showinfo("No Selected Files", "There are no selected files to download.")
            return
        
        # Prompt user to select the destination folder where "Selected Files" will be created
        destination_parent = filedialog.askdirectory(title="Select Destination for 'Selected Files' Folder")
        if not destination_parent:
            return  # User cancelled the dialog
        
        # Define the "Selected Files" subfolder path
        selected_folder_name = "Selected Files"
        destination_folder = os.path.join(destination_parent, selected_folder_name)
        
        # Check if "Selected Files" folder already exists
        if not os.path.exists(destination_folder):
            try:
                os.makedirs(destination_folder)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create 'Selected Files' folder: {str(e)}")
                return
        else:
            # Ask the user if they want to overwrite existing files
            overwrite = messagebox.askyesno("Folder Exists", f"The folder '{selected_folder_name}' already exists in the selected destination.\nDo you want to overwrite its contents?")
            if overwrite:
                # Optionally, clear existing contents
                try:
                    for filename in os.listdir(destination_folder):
                        file_path = os.path.join(destination_folder, filename)
                        if os.path.isfile(file_path):
                            os.remove(file_path)
                        elif os.path.isdir(file_path):
                            shutil.rmtree(file_path)
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to clear existing 'Selected Files' folder: {str(e)}")
                    return
            else:
                return  # User chose not to overwrite; cancel the operation
        
        # Start copying in a separate thread to keep GUI responsive
        copy_thread = threading.Thread(target=self.copy_selected_files, args=(destination_folder,))
        copy_thread.start()

    
    def copy_flagged_files(self, destination_folder):
        copied_count = 0
        failed_files = []
        
        for filepath in self.flagged_files:
            try:
                shutil.copy2(filepath, destination_folder)
                copied_count += 1
            except Exception as e:
                failed_files.append((filepath, str(e)))
        
        # Notify the user upon completion
        if failed_files:
            error_messages = "\n".join([f"{os.path.basename(f[0])}: {f[1]}" for f in failed_files])
            message = f"Copied {copied_count} flagged files successfully.\nFailed to copy {len(failed_files)} files:\n{error_messages}"
            messagebox.showwarning("Download Completed with Errors", message)
        else:
            message = f"All {copied_count} flagged files were copied successfully to '{destination_folder}'."
            messagebox.showinfo("Download Completed", message)

    def copy_selected_files(self, destination_folder):
        copied_count = 0
        failed_files = []
        
        for filepath in self.selected_files:
            try:
                shutil.copy2(filepath, destination_folder)
                copied_count += 1
            except Exception as e:
                failed_files.append((filepath, str(e)))
        
        # Notify the user upon completion
        if failed_files:
            error_messages = "\n".join([f"{os.path.basename(f[0])}: {f[1]}" for f in failed_files])
            message = f"Copied {copied_count} selected files successfully.\nFailed to copy {len(failed_files)} files:\n{error_messages}"
            messagebox.showwarning("Download Completed with Errors", message)
        else:
            message = f"All {copied_count} selected files were copied successfully to '{destination_folder}'."
            messagebox.showinfo("Download Completed", message)

        
    def open_file(self, filepath):
        try:
            if sys.platform.startswith('darwin'):
                subprocess.call(['open', filepath])
            elif os.name == 'nt':
                os.startfile(filepath)
            elif os.name == 'posix':
                subprocess.call(['xdg-open', filepath])
            else:
                messagebox.showerror("Unsupported OS", "Cannot open log file on this operating system.")
        except FileNotFoundError:
            messagebox.showerror("Error", f"The file '{filepath}' does not exist.")
        except PermissionError:
            messagebox.showerror("Error", f"Permission denied while trying to open '{filepath}'.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open log file: {str(e)}")
